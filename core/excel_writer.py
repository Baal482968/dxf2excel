"""
Excel 輸出相關功能模組 - 增強版
支援圖片嵌入和文字描述的混合模式
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
import tempfile
import os
import re

# 圖形相關模組
try:
    from utils.graphics.manager import GraphicsManager
    print("✅ 圖形管理器初始化成功")
except ImportError:
    GraphicsManager = None
    print("⚠️ 圖形管理器初始化失敗")

class ExcelWriter:
    """Excel 檔案寫入器 - 增強版"""
    
    def __init__(self, image_mode="mixed"):
        """
        初始化 Excel 寫入器
        
        Args:
            image_mode: 圖片處理模式
                - "image": 僅嵌入圖片
                - "text": 僅使用文字描述
                - "mixed": 圖片+文字描述（推薦）
                - "auto": 自動檢測並選擇最佳模式
        """
        self.workbook = None
        self.worksheet = None
        self.temp_files = []  # 暫存圖片檔案列表
        self.image_mode = image_mode
        
        # 圖形管理器初始化
        if GraphicsManager:
            try:
                self.graphics_manager = GraphicsManager()
                self.graphics_available = True
                print("✅ 圖形管理器初始化成功")
            except Exception as e:
                print(f"⚠️ 圖形管理器初始化失敗: {e}")
                self.graphics_available = False
        else:
            self.graphics_available = False
            print("⚠️ 圖形管理器不可用")
        
        # 根據可用性調整模式
        if self.image_mode == "auto":
            if self.graphics_available:
                self.image_mode = "mixed"
            else:
                self.image_mode = "text"
                print("🔄 自動切換到文字模式")
        
        # 定義樣式
        self.styles = {
            'header_font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'normal_font': Font(name='Calibri', size=14),
            'small_font': Font(name='Calibri', size=12),
            'title_font': Font(name='Calibri', size=16, bold=True),
            'description_font': Font(name='Consolas', size=12),  # 等寬字體用於圖示描述
            'header_fill': PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid'),
            'light_fill': PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'thick_border': Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
        }
    
    def create_workbook(self):
        """創建新的工作簿"""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "鋼筋計料表"
    
    def save_workbook(self, file_path):
        """儲存工作簿，並在儲存後清理暫存檔案"""
        if self.workbook:
            try:
                # 檢查保存前的圖片狀態
                if hasattr(self, 'worksheet') and self.worksheet:
                    print(f"🔍 保存前工作表圖片數量: {len(self.worksheet._images)}")
                    if hasattr(self.worksheet, '_images') and self.worksheet._images:
                        for i, img in enumerate(self.worksheet._images):
                            print(f"   圖片 {i+1}: {img}")
                
                self.workbook.save(file_path)
                print(f"✅ Excel 檔案已儲存: {file_path}")
            except Exception as e:
                print(f"❌ Excel 儲存失敗: {e}")
                raise
        
        # 清理暫存圖檔
        self._cleanup_temp_files()
    
    def _cleanup_temp_files(self):
        """清理暫存檔案"""
        if hasattr(self, 'temp_files'):
            for temp_file in self.temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        print(f"🗑️ 已清理暫存檔案: {temp_file}")
                except Exception as e:
                    print(f"⚠️ 清理暫存檔案失敗: {e}")
            self.temp_files.clear()
    
    def write_header(self, start_row=2):
        """寫入表頭，可指定起始 row"""
        headers = [
            "編號", "號數", "A(cm)", "B(cm)", "C(cm)", "D(cm)", "E(cm)", "F(cm)", "G(cm)", 
            "圖示", "長度(cm)", "數量", "重量(kg)", "備註", "讀取CAD文字"
        ]
        column_widths = [8, 8, 8, 8, 8, 8, 8, 8, 8, 20, 12, 8, 12, 20, 45]

        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=start_row, column=col)
            cell.value = header
            cell.font = self.styles['header_font']
            cell.fill = self.styles['header_fill']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.styles['border']
        for col, width in enumerate(column_widths, 1):
            self.worksheet.column_dimensions[get_column_letter(col)].width = width
    
    def write_title(self, title, subtitle=None):
        """寫入標題和副標題"""
        # 主標題
        self.worksheet.merge_cells('A1:O1')
        cell = self.worksheet.cell(row=1, column=1)
        cell.value = title
        cell.font = self.styles['title_font']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.worksheet.row_dimensions[1].height = 30
        
        # 副標題（如果提供）
        if subtitle:
            self.worksheet.merge_cells('A2:O2')
            cell = self.worksheet.cell(row=2, column=1)
            cell.value = subtitle
            cell.font = self.styles['normal_font']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            self.worksheet.row_dimensions[2].height = 20
            
            # 調整表頭行號
            return 3
        
        return 2
    
    # 圖片處理方法已移除，改為使用 assets/materials/ 資料夾中的圖示檔案
    
    def _get_rebar_segments(self, rebar):
        """從鋼筋資料中提取分段長度"""
        # 直接檢查 segments 欄位
        if 'segments' in rebar and isinstance(rebar['segments'], list) and rebar['segments']:
            return rebar['segments']
        # 如果沒有 segments，嘗試其他欄位
        segments = []
        segment_keys = ['lengths', 'A', 'B', 'C', 'D', 'E']
        for key in segment_keys:
            if key in rebar and rebar[key] is not None:
                if key == 'lengths' and isinstance(rebar[key], list):
                    segments = rebar[key]
                    break
                elif key in ['A', 'B', 'C', 'D', 'E']:
                    if key == 'A' and segments == []:
                        segments = []
                    if rebar[key] > 0:
                        segments.append(rebar[key])
        # 如果還是沒有分段資料，使用總長度
        if not segments and 'length' in rebar and rebar['length'] > 0:
            segments = [rebar['length']]
        return segments

    def _generate_rebar_visual(self, rebar):
        """
        生成鋼筋視覺表示（圖片或文字描述）
        """
        rebar_number = rebar.get('raw_text', rebar.get('rebar_number', '#4'))
        segments = self._get_rebar_segments(rebar)
        angles = rebar.get('angles', None)
        shape_type = rebar.get('type', None) # 取得箍筋類型
        
        # 針對箍筋，使用解析後的號數，而不是原始文字
        if shape_type and '箍' in shape_type:
            rebar_id = rebar.get('rebar_number', '#4')
        else:
            rebar_id = rebar_number

        # 檢查是否為 type10 鋼筋
        if shape_type == 'type10' and self.graphics_available:
            try:
                # 生成 type10 鋼筋圖片
                length = segments[0] if segments else 0
                image = self.graphics_manager.generate_type10_rebar_image(length, rebar_id)
                
                if image:
                    # 保存到臨時檔案
                    import tempfile
                    temp_img_path = tempfile.mktemp(suffix='.png')
                    image.save(temp_img_path)
                    self.temp_files.append(temp_img_path)
                    
                    print(f"🔍 生成 type10 鋼筋圖片: {temp_img_path}")
                    return temp_img_path
                    
            except Exception as e:
                print(f"⚠️ 生成 type10 鋼筋圖片失敗: {e}")
        
        # 生成文字描述
        if len(segments) == 1:
            text_description = f"直鋼筋 {rebar_id}\n長度: {int(segments[0])}cm"
        elif len(segments) == 2:
            text_description = f"L型鋼筋 {rebar_id}\n{int(segments[0])} + {int(segments[1])}cm"
        elif len(segments) == 3:
            text_description = f"U型鋼筋 {rebar_id}\n{int(segments[0])} + {int(segments[1])} + {int(segments[2])}cm"
        else:
            text_description = f"複雜鋼筋 {rebar_id}\n{' + '.join(str(int(s)) for s in segments)}cm"
            
        return text_description

    def write_rebar_data(self, rebar_data, start_row=3):
        """
        將鋼筋資料寫入工作表，包含圖示和詳細描述
        
        Args:
            rebar_data: 包含鋼筋資訊的字典列表
            start_row: 起始寫入行號
        
        Returns:
            int: 下一個可用行號
        """
        current_row = start_row
        for idx, rebar in enumerate(rebar_data, 1):
            # 基本資料
            self.worksheet.cell(row=current_row, column=1).value = idx
            self.worksheet.cell(row=current_row, column=2).value = rebar.get('rebar_number', '')
            
            # 確保 rebar 資料包含 segments
            if 'segments' not in rebar or not rebar['segments']:
                rebar['segments'] = self._get_rebar_segments(rebar)

            # 寫入 A-G 欄位
            segments = rebar.get('segments', [])
            for i, segment in enumerate(segments):
                if i < 7: # 最多寫入 7 個分段
                    self.worksheet.cell(row=current_row, column=3 + i).value = segment
            
            # 生成鋼筋視覺表示
            visual_info = self._generate_rebar_visual(rebar)
            
            # 圖示欄處理
            diagram_cell = self.worksheet.cell(row=current_row, column=10) # 圖示在第10欄
            
            # 檢查是否為圖片路徑
            if isinstance(visual_info, str) and os.path.exists(visual_info) and self.image_mode in ['image', 'mixed']:
                # 插入圖片
                try:
                    print(f"🔍 嘗試插入圖片: {visual_info}")
                    img = ExcelImage(visual_info)
                    # 調整圖片大小
                    img.width = 120
                    img.height = 80
                    
                    # 先清空圖示欄的文字內容
                    diagram_cell.value = ""
                    
                    # 使用 Claude 建議的正確語法
                    self.worksheet.add_image(img, f'J{current_row}')
                    
                    print(f"✅ 圖片插入成功到儲存格 J{current_row}")
                    
                    # 檢查圖片是否真的被添加
                    print(f"🔍 工作表圖片數量: {len(self.worksheet._images)}")
                    
                    # 再次確保圖示欄是空的
                    diagram_cell.value = ""
                    
                    # 調整行高以容納圖片
                    self.worksheet.row_dimensions[current_row].height = 80
                    
                except Exception as e:
                    print(f"⚠️ 圖片插入失敗: {e}")
                    # 如果圖片插入失敗，使用文字描述
                    diagram_cell.value = visual_info
                    self.worksheet.row_dimensions[current_row].height = 60
            else:
                # 使用文字描述
                diagram_cell.value = visual_info
                self.worksheet.row_dimensions[current_row].height = 60
            
            # 其他資料欄位
            self.worksheet.cell(row=current_row, column=11).value = round(rebar.get('length', 0), 1)
            self.worksheet.cell(row=current_row, column=12).value = rebar.get('count', 1)
            self.worksheet.cell(row=current_row, column=13).value = round(rebar.get('weight', 0), 1)
            self.worksheet.cell(row=current_row, column=14).value = rebar.get('note', '')
            self.worksheet.cell(row=current_row, column=15).value = rebar.get('raw_text', '')
            
            # 設定儲存格樣式
            for col in range(1, 16):
                cell = self.worksheet.cell(row=current_row, column=col)
                if col != 10:  # 圖示欄已單獨處理
                    cell.font = self.styles['normal_font']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.styles['border']
            
            current_row += 1
            
        return current_row
    
    def write_summary(self, rebar_data, start_row):
        """寫入統計摘要"""
        if not rebar_data:
            return start_row
        
        # 計算統計資料
        total_count = sum(rebar.get('count', 1) for rebar in rebar_data)
        total_weight = sum(rebar.get('weight', 0) for rebar in rebar_data)
        total_length = sum(rebar.get('length', 0) * rebar.get('count', 1) for rebar in rebar_data)
        
        # 鋼筋類型統計
        rebar_types = {}
        for rebar in rebar_data:
            rebar_num = rebar.get('rebar_number', '')
            if rebar_num not in rebar_types:
                rebar_types[rebar_num] = {'count': 0, 'weight': 0}
            rebar_types[rebar_num]['count'] += rebar.get('count', 1)
            rebar_types[rebar_num]['weight'] += rebar.get('weight', 0)
        
        # 寫入摘要標題
        summary_row = start_row + 1
        self.worksheet.merge_cells(f'A{summary_row}:O{summary_row}')
        cell = self.worksheet.cell(row=summary_row, column=1)
        cell.value = "統計摘要"
        cell.font = Font(name='Calibri', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
        cell.border = self.styles['thick_border']
        
        # 總計資料
        summary_row += 1
        summary_data = [
            ("總數量", f"{total_count} 支"),
            ("總重量", f"{total_weight:.1f} kg"),
            ("總長度", f"{total_length:.1f} cm"),
            ("鋼筋類型", f"{len(rebar_types)} 種")
        ]
        
        for i, (label, value) in enumerate(summary_data):
            label_cell = self.worksheet.cell(row=summary_row, column=i*2+1)
            value_cell = self.worksheet.cell(row=summary_row, column=i*2+2)
            
            label_cell.value = label
            value_cell.value = value
            
            label_cell.font = Font(name='Calibri', size=10, bold=True)
            value_cell.font = Font(name='Calibri', size=10)
            
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            value_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            label_cell.border = self.styles['border']
            value_cell.border = self.styles['border']
        
        return summary_row + 1
    
    def write_footer(self, row):
        """寫入頁尾"""
        # 生成時間
        self.worksheet.merge_cells(f'A{row}:O{row}')
        cell = self.worksheet.cell(row=row, column=1)
        
        # 模式資訊 - 圖示功能暫時停用
        cell.value = (f"生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
                     f"圖示功能：暫時停用，等 assets/materials 圖片準備好時再實作")
        cell.font = self.styles['small_font']
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = self.styles['border']
    
    def format_worksheet(self):
        """格式化工作表"""
        if not self.worksheet:
            return
        
        # 設定列印區域
        if self.worksheet.max_row > 0 and self.worksheet.max_column > 0:
            self.worksheet.print_area = f'A1:{get_column_letter(self.worksheet.max_column)}{self.worksheet.max_row}'
        
        # 設定頁面方向為橫向
        self.worksheet.page_setup.orientation = self.worksheet.ORIENTATION_LANDSCAPE
        
        # 設定頁面邊距
        self.worksheet.page_margins.left = 0.5
        self.worksheet.page_margins.right = 0.5
        self.worksheet.page_margins.top = 0.5
        self.worksheet.page_margins.bottom = 0.5
        
        # 設定頁首頁尾
        self.worksheet.oddHeader.center.text = "鋼筋計料表"
        self.worksheet.oddFooter.right.text = "第 &P 頁，共 &N 頁"
        
        # 凍結窗格（凍結表頭）
        self.worksheet.freeze_panes = 'A3'
        
        # 設定自動篩選
        if self.worksheet.max_row > 2:
            self.worksheet.auto_filter.ref = f'A2:{get_column_letter(self.worksheet.max_column)}{self.worksheet.max_row}'
    
    def write_multi_sheet_rebar_data(self, grouped_data, main_title="鋼筋計料表"):
        """依據分組資料寫入多個 sheet，每個分組一張表"""
        if not self.workbook:
            self.create_workbook()
        first = True
        for sheet_name, rebar_list in grouped_data.items():
            if first:
                ws = self.worksheet
                ws.title = sheet_name if sheet_name else "料表"
                first = False
            else:
                ws = self.workbook.create_sheet(title=sheet_name if sheet_name else "料表")
            self.worksheet = ws
            header_row = self.write_title(main_title, subtitle=sheet_name)
            self.write_header(start_row=header_row)
            next_row = self.write_rebar_data(rebar_list, start_row=header_row + 1)
            summary_row = self.write_summary(rebar_list, next_row)
            self.write_footer(summary_row + 1)
            self.format_worksheet()


# 便利函數
def create_excel_writer(mode="auto"):
    """
    創建 Excel 寫入器的便利函數
    
    Args:
        mode: 圖片處理模式
            - "auto": 自動檢測
            - "mixed": 圖文混合（推薦）
            - "image": 僅圖片
            - "text": 僅文字
    
    Returns:
        ExcelWriter: Excel 寫入器實例
    """
    return ExcelWriter(image_mode=mode)


def quick_generate_excel(rebar_data, output_path, title="鋼筋計料表", mode="auto"):
    """
    快速生成 Excel 檔案的便利函數
    
    Args:
        rebar_data: 鋼筋資料列表
        output_path: 輸出檔案路徑
        title: 表格標題
        mode: 圖片處理模式
    
    Returns:
        bool: 生成成功返回 True
    """
    try:
        writer = ExcelWriter(image_mode=mode)
        writer.create_workbook()
        
        # 寫入標題
        header_row = writer.write_title(title)
        
        # 寫入表頭
        writer.write_header()
        
        # 寫入資料
        next_row = writer.write_rebar_data(rebar_data, header_row + 1)
        
        # 寫入統計摘要
        summary_row = writer.write_summary(rebar_data, next_row)
        
        # 寫入頁尾
        writer.write_footer(summary_row + 1)
        
        # 格式化
        writer.format_worksheet()
        
        # 儲存
        writer.save_workbook(output_path)
        
        return True
        
    except Exception as e:
        print(f"❌ Excel 生成失敗: {e}")
        return False


# 測試函數
def test_excel_writer():
    """測試 Excel 寫入器功能"""
    
    # 模擬鋼筋資料
    test_data = [
        {
            'rebar_number': '#4',
            'length': 300,
            'count': 10,
            'weight': 12.5,
            'note': '主筋',
            'raw_text': '#4-300x10',
            'segments': [300]
        },
        {
            'rebar_number': '#5', 
            'length': 350,
            'count': 8,
            'weight': 18.7,
            'note': 'L型箍筋',
            'raw_text': '#5-150+200x8',
            'A': 150,
            'B': 200
        },
        {
            'rebar_number': '#6',
            'length': 470,
            'count': 6,
            'weight': 25.3,
            'note': 'U型箍筋',
            'raw_text': '#6-120+230+120x6',
            'segments': [120, 230, 120]
        }
    ]
    
    print("🔄 測試 Excel 寫入器")
    print("=" * 40)
    
    # 測試不同模式
    modes = ["mixed", "text", "image"]
    
    for mode in modes:
        try:
            output_file = f"test_rebar_{mode}.xlsx"
            print(f"\n📝 測試 {mode} 模式...")
            
            success = quick_generate_excel(
                test_data, 
                output_file, 
                f"鋼筋計料表 - {mode.upper()}模式測試",
                mode
            )
            
            if success:
                print(f"✅ {mode} 模式測試成功: {output_file}")
            else:
                print(f"❌ {mode} 模式測試失敗")
                
        except Exception as e:
            print(f"❌ {mode} 模式測試錯誤: {e}")
    
    print("\n🎉 測試完成")


if __name__ == "__main__":
    """當檔案被直接執行時進行測試"""
    test_excel_writer()