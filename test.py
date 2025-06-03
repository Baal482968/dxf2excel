"""
CAD 鋼筋計料轉換工具 Pro v2.1
=====================================

功能說明:
- 讀取 DXF 格式的 CAD 檔案
- 自動識別鋼筋文字標記並提取數據
- 生成專業的鋼筋彎曲圖示
- 輸出格式化的 Excel 鋼筋計料表
- 支援多種鋼筋規格和材料等級

作者: BaalWu
版本: v2.1 Professional Edition
建立時間: 2025-06-03
最後更新: 2025-06-03
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import ezdxf
import math
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import threading
import time
import re


class ModernCADtoExcelConverter:
    """
    現代化 CAD 到 Excel 轉換器主類別
    
    這個類別負責:
    1. 建立現代化的使用者介面
    2. 處理 CAD 檔案讀取和分析
    3. 生成專業的鋼筋圖示
    4. 輸出格式化的 Excel 報表
    """
    
    def __init__(self, root):
        """
        初始化轉換器
        
        Args:
            root: tkinter 主視窗物件
        
        初始化內容:
        - 設定視窗基本屬性
        - 定義配色主題
        - 初始化材質和鋼筋數據
        - 檢查圖形繪製功能
        - 建立使用者介面
        """
        try:
            # === 基本視窗設定 ===
            self.root = root
            self.root.title("CAD 鋼筋計料轉換工具 Pro v2.1")
            self.root.geometry("900x750")
            self.root.resizable(True, True)
            
            # === 現代化配色主題設定 ===
            self.colors = {
                'primary': '#4A90E2',      # 主要藍色 - 用於重要按鈕和標題
                'secondary': '#E24A90',    # 次要紫色 - 用於強調元素
                'success': '#F5A623',      # 成功橙色 - 用於成功狀態顯示
                'background': '#1E1E1E',   # 背景色（深灰） - 主要背景
                'surface': '#2D2D2D',      # 表面色（稍淺的深灰） - 卡片背景
                'text_primary': '#FFFFFF', # 主要文字（白色） - 主要文字顏色
                'text_secondary': '#B0B0B0', # 次要文字（淺灰） - 提示文字顏色
                'border': '#404040',       # 邊框色（中灰） - 元件邊框
                'accent': '#3D3D3D'        # 強調色（深灰） - 輔助元素
            }
            
            # 設定根視窗背景
            self.root.configure(bg=self.colors['background'])
            
            # === 材質密度設定 ===
            # 單位: kg/m³ - 用於重量計算
            self.material_density = {
                "鋼筋": 7850,      # 一般結構鋼筋密度
                "鋁": 2700,        # 鋁合金密度
                "銅": 8960,        # 銅材密度
                "不鏽鋼": 8000     # 不鏽鋼密度
            }
            
            # 預設材質類型
            self.default_material = "鋼筋"
            
            # === 鋼筋單位重量設定 ===
            # 單位: kg/m - 根據台灣 CNS 標準
            self.rebar_unit_weight = {
                "#2": 0.249,   # D6 - 直徑 6mm
                "#3": 0.561,   # D10 - 直徑 10mm
                "#4": 0.996,   # D13 - 直徑 13mm
                "#5": 1.552,   # D16 - 直徑 16mm
                "#6": 2.235,   # D19 - 直徑 19mm
                "#7": 3.042,   # D22 - 直徑 22mm
                "#8": 3.973,   # D25 - 直徑 25mm
                "#9": 5.026,   # D29 - 直徑 29mm
                "#10": 6.404,  # D32 - 直徑 32mm
                "#11": 7.906,  # D36 - 直徑 36mm
                "#12": 11.38,  # D43 - 直徑 43mm
                "#13": 13.87,  # D50 - 直徑 50mm
                "#14": 14.59,  # D57 - 直徑 57mm
                "#15": 20.24,  # 非標準規格
                "#16": 25.00,  # 非標準規格
                "#17": 31.20,  # 非標準規格
                "#18": 39.70   # 非標準規格
            }
            
            # === 標準鋼筋彎曲半徑設定 ===
            # 單位: 倍數（相對於鋼筋直徑）
            # 根據建築技術規則規定
            self.bend_radius = {
                '#2': 3, '#3': 3, '#4': 4, '#5': 5, '#6': 6, '#7': 7, '#8': 8,
                '#9': 9, '#10': 10, '#11': 11, '#12': 12, '#13': 13, '#14': 14,
                '#15': 15, '#16': 16, '#17': 17, '#18': 18
            }
            
            # === 處理進度追蹤變數 ===
            self.current_step = 0          # 目前處理步驟
            self.total_steps = 0           # 總步驟數
            self.step_descriptions = {}    # 各步驟描述
            self.processing_start_time = 0 # 處理開始時間
            
            # === 檢查圖形繪製功能是否可用 ===
            self.graphics_available = self.check_graphics_dependencies()
            
            # === 建立使用者介面 ===
            self.setup_modern_styles()  # 設定現代化樣式
            self.setup_ui()             # 建立使用者介面
            
        except Exception as e:
            print(f"初始化錯誤: {str(e)}")
            messagebox.showerror("錯誤", f"程式初始化時發生錯誤:\n{str(e)}")
    
    def check_graphics_dependencies(self):
        """
        檢查圖形繪製所需的套件是否已安裝
        
        Returns:
            bool: True 表示圖形功能可用，False 表示不可用
        
        檢查套件:
        - matplotlib: 用於繪製專業圖形
        - numpy: 數值計算支援
        - PIL (Pillow): 圖像處理
        """
        try:
            # 嘗試匯入所需套件
            import matplotlib
            import numpy
            from PIL import Image
            
            # 設定臨時日誌函數（稍後會被覆蓋）
            self.log_message = lambda x: print(x)
            return True
            
        except ImportError as e:
            # 檢查具體缺少哪些套件
            missing_packages = []
            
            try:
                import matplotlib
            except ImportError:
                missing_packages.append("matplotlib")
            
            try:
                import numpy
            except ImportError:
                missing_packages.append("numpy")
            
            try:
                from PIL import Image
            except ImportError:
                missing_packages.append("pillow")
            
            # 顯示缺少的套件信息
            print(f"⚠️ 缺少圖形繪製套件: {', '.join(missing_packages)}")
            print("請執行: pip install matplotlib numpy pillow")
            return False
    
    def install_graphics_dependencies(self):
        """
        嘗試自動安裝圖形繪製套件
        
        Returns:
            bool: 安裝成功返回 True，失敗返回 False
        
        安裝套件:
        - matplotlib: 圖形繪製
        - numpy: 數值計算
        - pillow: 圖像處理
        """
        try:
            import subprocess
            import sys
            
            packages = ["matplotlib", "numpy", "pillow"]
            self.log_message("🔄 正在安裝圖形繪製套件...")
            
            # 逐一檢查並安裝套件
            for package in packages:
                try:
                    # 檢查套件是否已安裝
                    __import__(package if package != "pillow" else "PIL")
                except ImportError:
                    # 套件未安裝，執行安裝
                    self.log_message(f"正在安裝 {package}...")
                    subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                    self.log_message(f"✅ {package} 安裝成功")
            
            # 更新圖形功能狀態
            self.graphics_available = True
            self.log_message("🎨 圖形繪製功能已啟用")
            return True
            
        except Exception as e:
            self.log_message(f"❌ 自動安裝失敗: {str(e)}")
            self.log_message("請手動執行: pip install matplotlib numpy pillow")
            return False
    
    # ====== 圖形化鋼筋繪製方法 ======
    
    def enhanced_draw_rebar_diagram(self, segments, rebar_number="#4"):
        """
        增強版鋼筋圖示生成方法
        
        Args:
            segments (list): 鋼筋分段長度列表
            rebar_number (str): 鋼筋編號，例如 "#4"
        
        Returns:
            str: 生成的鋼筋圖示文字
        
        功能:
        - 根據分段數量選擇不同的圖示樣式
        - 支援直鋼筋、L型、U型和複雜多段鋼筋
        - 當圖形功能不可用時自動降級為 ASCII 圖示
        """
        # 檢查圖形功能是否可用
        if not self.graphics_available:
            return self.draw_ascii_rebar(segments)
        
        try:
            # 檢查分段資料是否有效
            if not segments:
                return "無分段資料"
            
            # 根據分段數量選擇不同的繪製方法
            if len(segments) == 1:
                # 單段 = 直鋼筋
                return self._draw_straight_rebar_diagram(segments[0], rebar_number)
            elif len(segments) == 2:
                # 兩段 = L型鋼筋
                return self._draw_l_shaped_diagram(segments[0], segments[1], rebar_number)
            elif len(segments) == 3:
                # 三段 = U型鋼筋
                return self._draw_u_shaped_diagram(segments[0], segments[1], segments[2], rebar_number)
            else:
                # 多段 = 複雜彎曲鋼筋
                return self._draw_complex_rebar_diagram(segments, rebar_number)
                
        except Exception as e:
            # 發生錯誤時使用簡化圖示
            self.log_message(f"⚠️ 圖形生成錯誤: {str(e)}，使用簡化圖示")
            return self.draw_ascii_rebar(segments)
    
    def _draw_straight_rebar_diagram(self, length, rebar_number):
        """
        繪製直鋼筋的專業圖示
        
        Args:
            length (float): 鋼筋長度（公分）
            rebar_number (str): 鋼筋編號
        
        Returns:
            str: 直鋼筋圖示文字
        
        圖示包含:
        - 鋼筋編號和長度
        - 材料等級
        - 簡單的水平線條表示
        """
        try:
            # 格式化長度為整數字串
            l_str = str(int(length))
            # 取得材料等級
            material_grade = self._get_material_grade(rebar_number)
            
            # 建立圖示文字
            lines = []
            lines.append(f"直鋼筋 {rebar_number}")
            lines.append(f"長度: {l_str}cm")
            lines.append(f"等級: {material_grade}")
            lines.append("├" + "─" * (len(l_str) + 8) + "┤")
            lines.append(f"  {l_str}cm")
            lines.append("└" + "─" * (len(l_str) + 8) + "┘")
            
            return "\n".join(lines)
            
        except Exception:
            # 簡化版本（錯誤時使用）
            return f"直鋼筋 {rebar_number}: {int(length)}cm"

    def _draw_l_shaped_diagram(self, length1, length2, rebar_number):
        """
        繪製 L 型鋼筋專業圖示
        
        Args:
            length1 (float): 第一段長度（公分）
            length2 (float): 第二段長度（公分）
            rebar_number (str): 鋼筋編號
        
        Returns:
            str: L 型鋼筋圖示文字
        
        圖示特點:
        - 顯示兩段長度
        - 包含彎曲半徑資訊
        - 使用方框字元模擬 L 型形狀
        """
        try:
            # 格式化長度
            l1_str = str(int(length1))
            l2_str = str(int(length2))
            # 取得材料等級和彎曲半徑
            material_grade = self._get_material_grade(rebar_number)
            bend_r = self.bend_radius.get(rebar_number, 5)
            
            # 建立 L 型圖示
            lines = []
            lines.append(f"L型鋼筋 {rebar_number}")
            lines.append(f"等級: {material_grade}")
            lines.append(f"彎曲半徑: R{bend_r}cm")
            lines.append("")
            lines.append(f"    {l1_str}cm")
            lines.append("┌" + "─" * (max(len(l1_str), 8) + 2) + "┐")
            lines.append("│" + " " * (max(len(l1_str), 8) + 2) + "│")
            lines.append("│" + " " * (max(len(l1_str), 8) + 2) + "│")
            lines.append("│" + " " * (max(len(l1_str), 8) + 2) + f"│ {l2_str}cm")
            lines.append("│" + " " * (max(len(l1_str), 8) + 2) + "│")
            lines.append("│" + " " * (max(len(l1_str), 8) + 2) + "│")
            lines.append("└" + "─" * (max(len(l1_str), 8) + 2) + "┘")
            
            return "\n".join(lines)
            
        except Exception:
            # 簡化版本
            return f"L型 {rebar_number}: {int(length1)}+{int(length2)}cm"

    def _draw_u_shaped_diagram(self, length1, length2, length3, rebar_number):
        """
        繪製 U 型鋼筋專業圖示
        
        Args:
            length1 (float): 第一段長度（公分）
            length2 (float): 第二段長度（公分）- 底部橫段
            length3 (float): 第三段長度（公分）
            rebar_number (str): 鋼筋編號
        
        Returns:
            str: U 型鋼筋圖示文字
        
        圖示特點:
        - 顯示三段長度
        - 模擬 U 型結構
        - 包含材料等級和彎曲半徑
        """
        try:
            # 格式化長度
            l1_str = str(int(length1))
            l2_str = str(int(length2))
            l3_str = str(int(length3))
            # 取得相關資訊
            material_grade = self._get_material_grade(rebar_number)
            bend_r = self.bend_radius.get(rebar_number, 5)
            
            # 建立 U 型圖示
            lines = []
            lines.append(f"U型鋼筋 {rebar_number}")
            lines.append(f"等級: {material_grade}")
            lines.append(f"彎曲半徑: R{bend_r}cm")
            lines.append("")
            lines.append(f"{l2_str}cm     {l1_str}cm     {l3_str}cm")
            lines.append("│" + " " * (max(len(l1_str), 12) + 4) + "│")
            lines.append("│" + " " * (max(len(l1_str), 12) + 4) + "│")
            lines.append("│" + " " * (max(len(l1_str), 12) + 4) + "│")
            lines.append("│" + " " * (max(len(l1_str), 12) + 4) + "│")
            lines.append("└" + "─" * (max(len(l1_str), 12) + 4) + "┘")
            
            return "\n".join(lines)
            
        except Exception:
            # 簡化版本
            return f"U型 {rebar_number}: {int(length1)}+{int(length2)}+{int(length3)}cm"

    def _draw_complex_rebar_diagram(self, segments, rebar_number):
        """
        繪製複雜多段鋼筋專業圖示
        
        Args:
            segments (list): 分段長度列表
            rebar_number (str): 鋼筋編號
        
        Returns:
            str: 複雜鋼筋圖示文字
        
        功能:
        - 處理 4 段以上的複雜鋼筋
        - 根據段數選擇不同的圖案樣式
        - 顯示總長度和分段資訊
        """
        try:
            # 計算基本資訊
            total_length = sum(segments)
            segment_str = "+".join([str(int(s)) for s in segments])
            material_grade = self._get_material_grade(rebar_number)
            
            # 建立基本資訊
            lines = []
            lines.append(f"多段彎曲鋼筋 {rebar_number}")
            lines.append(f"等級: {material_grade}")
            lines.append(f"分段: {segment_str}cm")
            lines.append(f"總長: {int(total_length)}cm")
            lines.append("")
            
            # 根據段數創建不同的圖案樣式
            if len(segments) == 4:
                # 四段鋼筋圖案
                lines.extend([
                    "┌─────────┬─────────┐",
                    "│         │         │",
                    "│         │         │",
                    "│         └─────────┤",
                    "│                   │",
                    "└───────────────────┘"
                ])
            elif len(segments) == 5:
                # 五段鋼筋圖案
                lines.extend([
                    "┌─────┬─────┬─────┐",
                    "│     │     │     │",
                    "│     │     └─────┤",
                    "│     │           │",
                    "│     └───────────┤",
                    "│                 │",
                    "└─────────────────┘"
                ])
            else:
                # 通用多段圖案（6段以上）
                width = min(len(segments) * 4, 20)
                lines.extend([
                    "┌" + "─┬─" * min(len(segments)-1, 5) + "─┐",
                    "│" + " │ " * min(len(segments)-1, 5) + " │",
                    "│" + " └─" * min(len(segments)-1, 5) + " │",
                    "│" + " " * width + "│",
                    "└" + "─" * width + "┘"
                ])
            
            return "\n".join(lines)
            
        except Exception:
            # 簡化版本
            return f"多段 {rebar_number}: {'+'.join([str(int(s)) for s in segments])}cm"
    
    def _get_material_grade(self, rebar_number):
        """
        根據鋼筋編號獲取材料等級
        
        Args:
            rebar_number (str): 鋼筋編號，例如 "#4"
        
        Returns:
            str: 材料等級 (SD280/SD420/SD490)
        
        等級分類標準:
        - #2~#6: SD280 (一般結構用)
        - #7~#10: SD420 (高強度結構用)
        - #11以上: SD490 (特殊高強度)
        """
        if rebar_number.startswith('#'):
            try:
                # 提取編號數字
                num = int(rebar_number[1:])
                # 根據編號分級
                if num <= 6:
                    return "SD280"    # 一般結構用鋼筋
                elif num <= 10:
                    return "SD420"    # 高強度結構鋼筋
                else:
                    return "SD490"    # 特殊高強度鋼筋
            except:
                pass
        # 預設等級
        return "SD280"
    
    # ====== 保留原有的所有方法 ======
    
    def setup_modern_styles(self):
        """
        設定現代化使用者介面樣式
        
        功能:
        - 設定 ttk 樣式主題
        - 定義各種元件的外觀樣式
        - 配置顏色、字體、邊框等視覺效果
        
        樣式類別:
        - Modern.TFrame: 現代化框架樣式
        - Card.TFrame: 卡片式框架樣式
        - Header/Title/Body/Caption.TLabel: 各級文字樣式
        - Primary/Secondary.TButton: 主要和次要按鈕樣式
        """
        self.style = ttk.Style()
        
        # 設定基礎主題
        self.style.theme_use('clam')
        
        # === 框架樣式配置 ===
        self.style.configure("Modern.TFrame", 
                           background=self.colors['surface'],
                           relief='flat',
                           borderwidth=1)
        
        self.style.configure("Card.TFrame",
                           background=self.colors['surface'],
                           relief='solid',
                           borderwidth=1,
                           bordercolor=self.colors['border'])
        
        # === 文字標籤樣式配置 ===
        # 主標題樣式
        self.style.configure("Header.TLabel",
                           background=self.colors['surface'],
                           foreground=self.colors['primary'],
                           font=("Segoe UI", 18, "bold"))
        
        # 副標題樣式
        self.style.configure("Title.TLabel",
                           background=self.colors['surface'],
                           foreground=self.colors['text_primary'],
                           font=("Segoe UI", 12, "bold"))
        
        # 內文樣式
        self.style.configure("Body.TLabel",
                           background=self.colors['surface'],
                           foreground=self.colors['text_primary'],
                           font=("Segoe UI", 10))
        
        # 說明文字樣式
        self.style.configure("Caption.TLabel",
                           background=self.colors['surface'],
                           foreground=self.colors['text_secondary'],
                           font=("Segoe UI", 9))
        
        # 成功狀態文字樣式
        self.style.configure("Success.TLabel",
                           background=self.colors['surface'],
                           foreground=self.colors['success'],
                           font=("Segoe UI", 10, "bold"))
        
        # === 按鈕樣式配置 ===
        # 主要按鈕樣式（重要操作）
        self.style.configure("Primary.TButton",
                           background='black',
                           foreground='#FFD700',  # 亮黃色文字
                           font=("Segoe UI", 14, "bold"),
                           relief='flat',
                           borderwidth=0,
                           focuscolor='none')
        
        # 次要按鈕樣式（一般操作）
        self.style.configure("Secondary.TButton",
                           background='black',
                           foreground='#FFD700',  # 亮黃色文字
                           font=("Segoe UI", 12),
                           relief='flat',
                           borderwidth=1,
                           focuscolor='none')
        
        # === 輸入框樣式配置 ===
        self.style.configure("Modern.TEntry",
                           fieldbackground='white',
                           borderwidth=1,
                           relief='solid',
                           font=("Segoe UI", 10))
        
        # === 進度條樣式配置 ===
        self.style.configure("Horizontal.TProgressbar",
                           background=self.colors['primary'],    # 進度條顏色
                           troughcolor=self.colors['accent'],    # 背景軌道顏色
                           borderwidth=0,
                           lightcolor=self.colors['primary'],
                           darkcolor=self.colors['primary'])
        
        # === 滑鼠懸停效果配置 ===
        # 主要按鈕懸停效果
        self.style.map("Primary.TButton",
                      background=[('active', '#333333')])  # 深灰色懸停
        
        # 次要按鈕懸停效果
        self.style.map("Secondary.TButton",
                      background=[('active', '#333333')])  # 深灰色懸停
    
    def create_card_frame(self, parent, title="", padding=20):
        """
        創建卡片式框架
        
        Args:
            parent: 父容器
            title (str): 卡片標題
            padding (int): 內邊距
        
        Returns:
            tk.Frame: 卡片內容框架
        
        功能:
        - 創建具有陰影效果的卡片式容器
        - 可選擇性添加標題
        - 自動設定內邊距和樣式
        """
        # === 創建陰影效果框架 ===
        shadow_frame = tk.Frame(parent, bg=self.colors['border'], height=2)
        shadow_frame.pack(fill=tk.X, pady=(5, 0), padx=2)
        
        # === 創建主卡片框架 ===
        card_frame = tk.Frame(parent, bg=self.colors['surface'], relief='solid', bd=1)
        card_frame.pack(fill=tk.X, pady=(0, 15), padx=5)
        
        # === 添加標題（如果提供） ===
        if title:
            title_frame = tk.Frame(card_frame, bg=self.colors['surface'])
            title_frame.pack(fill=tk.X, pady=(15, 5), padx=padding)
            
            title_label = tk.Label(title_frame, text=title,
                                 bg=self.colors['surface'],
                                 fg=self.colors['text_primary'],
                                 font=("Segoe UI", 12, "bold"))
            title_label.pack(side=tk.LEFT)
        
        # === 創建內容框架 ===
        content_frame = tk.Frame(card_frame, bg=self.colors['surface'])
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 15), padx=padding)
        
        return content_frame
    
    def create_material_button(self, parent, text, command, is_primary=True):
        """
        創建 Material Design 風格的按鈕
        
        Args:
            parent: 父容器
            text (str): 按鈕文字
            command: 按鈕點擊事件處理函數
            is_primary (bool): 是否為主要按鈕
        
        Returns:
            tuple: (按鈕物件, 陰影框架物件)
        
        功能:
        - 創建具有陰影效果的現代化按鈕
        - 支援主要和次要按鈕樣式
        - 自動添加滑鼠懸停效果
        """
        # === 創建外層框架用於陰影效果 ===
        shadow_frame = tk.Frame(parent, bg=self.colors['background'])
        
        # 設定按鈕文字顏色（深灰色以提高可讀性）
        text_color = '#222222'
        
        # === 創建按鈕 ===
        btn = tk.Button(shadow_frame, text=text,
                       command=command,
                       # 主要按鈕使用 Material Blue，次要按鈕使用深灰色
                       bg='#2196F3' if is_primary else '#424242',
                       fg=text_color,
                       # 懸停時的顏色
                       activebackground='#1976D2' if is_primary else '#616161',
                       activeforeground=text_color,
                       # 焦點顏色
                       highlightbackground='#2196F3' if is_primary else '#424242',
                       highlightcolor='#2196F3' if is_primary else '#424242',
                       font=("Segoe UI", 12, "bold"),
                       relief='flat',        # 扁平化設計
                       bd=0,                # 無邊框
                       cursor='hand2',      # 手型游標
                       padx=20,             # 水平內邊距
                       pady=10)             # 垂直內邊距
        btn.pack(fill=tk.BOTH, expand=True)
        
        # === 移除高亮邊框以達到圓角效果 ===
        btn.configure(highlightthickness=0)
        
        # === 添加滑鼠懸停效果 ===
        def on_enter(e):
            """滑鼠進入時的效果"""
            btn.configure(bg='#1976D2' if is_primary else '#616161', fg=text_color)
            shadow_frame.configure(bg='#1976D2' if is_primary else '#616161')
        
        def on_leave(e):
            """滑鼠離開時的效果"""
            btn.configure(bg='#2196F3' if is_primary else '#424242', fg=text_color)
            shadow_frame.configure(bg=self.colors['background'])
        
        # 綁定懸停事件
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        return btn, shadow_frame
    
    def setup_ui(self):
        """
        設定現代化使用者介面
        
        功能:
        - 建立完整的使用者介面佈局
        - 包含標題區域、輸入設定、控制按鈕、狀態顯示等
        - 設定鍵盤快捷鍵
        - 顯示初始狀態訊息
        
        介面結構:
        1. 標題區域 - 程式名稱和版本資訊
        2. 輸入檔案卡片 - CAD 檔案選擇
        3. 輸出檔案卡片 - Excel 檔案設定
        4. 控制按鈕卡片 - 執行和重置按鈕
        5. 處理狀態卡片 - 進度顯示和詳細日誌
        """
        # === 主容器設定 ===
        main_container = tk.Frame(self.root, bg=self.colors['background'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # === 標題區域 ===
        header_frame = tk.Frame(main_container, bg=self.colors['background'])
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        # 主標題
        title_label = tk.Label(header_frame,
                             text="🏗️ CAD 鋼筋計料轉換工具 Pro",
                             bg=self.colors['background'],
                             fg=self.colors['primary'],
                             font=("Segoe UI", 20, "bold"))
        title_label.pack()
        
        # 副標題 - 根據圖形功能狀態顯示不同訊息
        graphics_status = "✅ 圖形化功能已啟用" if self.graphics_available else "⚠️ 圖形化功能未啟用"
        subtitle_label = tk.Label(header_frame,
                                text=f"專業級 DXF 檔案鋼筋數據分析與 Excel 報表生成工具 | {graphics_status}",
                                bg=self.colors['background'],
                                fg=self.colors['text_secondary'],
                                font=("Segoe UI", 11))
        subtitle_label.pack(pady=(5, 0))
        
        # 版本標籤
        version_label = tk.Label(header_frame,
                               text="v2.1 Professional Edition with Graphics",
                               bg=self.colors['background'],
                               fg=self.colors['success'],
                               font=("Segoe UI", 9, "bold"))
        version_label.pack(pady=(5, 0))
        
        # === 圖形套件安裝按鈕（僅在功能未啟用時顯示） ===
        if not self.graphics_available:
            install_frame = tk.Frame(header_frame, bg=self.colors['background'])
            install_frame.pack(pady=(10, 0))
            
            self.install_button, self.install_shadow = self.create_material_button(
                install_frame, "🔧 安裝圖形繪製套件", self.install_graphics_dependencies, True)
        
        # === 輸入檔案設定卡片 ===
        input_card = self.create_card_frame(main_container, "📁 輸入檔案設定")
        
        # CAD 檔案選擇區域
        cad_section = tk.Frame(input_card, bg=self.colors['surface'])
        cad_section.pack(fill=tk.X, pady=(0, 15))
        
        # CAD 檔案標籤
        cad_label = tk.Label(cad_section, text="CAD 檔案",
                           bg=self.colors['surface'],
                           fg=self.colors['text_primary'],
                           font=("Segoe UI", 10, "bold"))
        cad_label.pack(anchor='w')
        
        # CAD 檔案輸入框架
        cad_input_frame = tk.Frame(cad_section, bg=self.colors['surface'])
        cad_input_frame.pack(fill=tk.X, pady=(5, 0))
        
        # CAD 檔案路徑變數和輸入框
        self.cad_path = tk.StringVar()
        self.cad_entry = tk.Entry(cad_input_frame, textvariable=self.cad_path,
                                font=("Segoe UI", 11),
                                bg=self.colors['surface'],
                                fg=self.colors['text_primary'],
                                insertbackground=self.colors['primary'],  # 游標顏色
                                relief='solid', bd=1,
                                selectbackground=self.colors['primary'],  # 選取背景色
                                selectforeground='white')                 # 選取文字色
        self.cad_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8)
        
        # CAD 檔案瀏覽按鈕
        self.browse_cad_button, self.browse_cad_shadow = self.create_material_button(
            cad_input_frame, "📂 瀏覽", self.browse_cad_file, True)
        self.browse_cad_shadow.pack(side=tk.RIGHT, padx=(10, 0), ipady=2)
        
        # 檔案資訊顯示標籤
        self.file_info_label = tk.Label(cad_section, text="",
                                      bg=self.colors['surface'],
                                      fg=self.colors['text_secondary'],
                                      font=("Segoe UI", 9))
        self.file_info_label.pack(anchor='w', pady=(5, 0))
        
        # === 輸出檔案設定卡片 ===
        output_card = self.create_card_frame(main_container, "💾 輸出檔案設定")
        
        # Excel 檔案選擇區域
        excel_section = tk.Frame(output_card, bg=self.colors['surface'])
        excel_section.pack(fill=tk.X)
        
        # Excel 檔案標籤
        excel_label = tk.Label(excel_section, text="Excel 輸出檔案",
                             bg=self.colors['surface'],
                             fg=self.colors['text_primary'],
                             font=("Segoe UI", 10, "bold"))
        excel_label.pack(anchor='w')
        
        # Excel 檔案輸入框架
        excel_input_frame = tk.Frame(excel_section, bg=self.colors['surface'])
        excel_input_frame.pack(fill=tk.X, pady=(5, 0))
        
        # Excel 檔案路徑變數和輸入框
        self.excel_path = tk.StringVar()
        excel_entry = tk.Entry(excel_input_frame, textvariable=self.excel_path,
                             font=("Segoe UI", 11),
                             bg=self.colors['surface'],
                             fg=self.colors['text_primary'],
                             insertbackground=self.colors['primary'],
                             relief='solid', bd=1,
                             selectbackground=self.colors['primary'],
                             selectforeground='white')
        excel_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8)
        
        # Excel 檔案瀏覽按鈕
        self.browse_excel_button, self.browse_excel_shadow = self.create_material_button(
            excel_input_frame, "💾 另存新檔", self.browse_excel_file, True)
        self.browse_excel_shadow.pack(side=tk.RIGHT, padx=(10, 0), ipady=2)
        
        # === 控制按鈕區域卡片 ===
        control_card = self.create_card_frame(main_container, "🎮 執行控制")
        
        button_frame = tk.Frame(control_card, bg=self.colors['surface'])
        button_frame.pack(fill=tk.X)
        
        # 主要轉換按鈕（根據圖形功能狀態顯示不同文字）
        button_text = "🚀 開始轉換 (含圖形)" if self.graphics_available else "🚀 開始轉換"
        self.convert_button, self.convert_shadow = self.create_material_button(
            button_frame, button_text, self.start_conversion, True)
        self.convert_shadow.pack(side=tk.LEFT, padx=(0, 10), ipady=2)
        
        # 重置按鈕
        self.reset_button, self.reset_shadow = self.create_material_button(
            button_frame, "🔄 重置", self.reset_form, False)
        self.reset_shadow.pack(side=tk.LEFT, padx=(0, 10), ipady=2)
        
        # 快捷鍵提示標籤
        shortcut_label = tk.Label(button_frame,
                                text="💡 快捷鍵: Ctrl+O(開檔) | Ctrl+S(存檔) | F5(轉換) | Esc(重置)",
                                bg=self.colors['surface'],
                                fg=self.colors['text_secondary'],
                                font=("Segoe UI", 9))
        shortcut_label.pack(side=tk.LEFT, padx=(10, 0))
        
        # === 處理狀態與進度卡片 ===
        status_card = self.create_card_frame(main_container, "📊 處理狀態與進度")
        
        # 進度資訊區域
        progress_info_frame = tk.Frame(status_card, bg=self.colors['surface'])
        progress_info_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 當前步驟顯示標籤
        self.current_step_label = tk.Label(progress_info_frame, text="⭐ 準備就緒",
                                         bg=self.colors['surface'],
                                         fg=self.colors['text_primary'],
                                         font=("Segoe UI", 11, "bold"))
        self.current_step_label.pack(side=tk.LEFT)
        
        # 時間和百分比顯示標籤
        self.time_label = tk.Label(progress_info_frame, text="",
                                 bg=self.colors['surface'],
                                 fg=self.colors['text_secondary'],
                                 font=("Segoe UI", 10))
        self.time_label.pack(side=tk.RIGHT)
        
        # 現代化進度條
        progress_frame = tk.Frame(status_card, bg=self.colors['surface'])
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress = ttk.Progressbar(progress_frame,
                                      orient="horizontal",
                                      mode="determinate",
                                      style="Horizontal.TProgressbar")
        self.progress.pack(fill=tk.X, ipady=8)
        
        # 詳細進度說明標籤
        self.detail_progress_label = tk.Label(status_card, text="",
                                            bg=self.colors['surface'],
                                            fg=self.colors['text_secondary'],
                                            font=("Segoe UI", 9))
        self.detail_progress_label.pack(fill=tk.X, pady=(0, 15))
        
        # === 狀態文字區域 ===
        text_container = tk.Frame(status_card, bg=self.colors['surface'])
        text_container.pack(fill=tk.BOTH, expand=True)
        
        # 狀態文字標題
        text_title = tk.Label(text_container, text="📝 詳細處理日誌",
                            bg=self.colors['surface'],
                            fg=self.colors['text_primary'],
                            font=("Segoe UI", 10, "bold"))
        text_title.pack(anchor='w', pady=(0, 5))
        
        # 文字框架（包含文字框和捲軸）
        text_frame = tk.Frame(text_container, bg=self.colors['surface'])
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        # 狀態文字框
        self.status_text = tk.Text(text_frame, height=10,
                                 bg=self.colors['surface'],
                                 fg=self.colors['text_primary'],
                                 font=("Consolas", 10),  # 等寬字體便於對齊
                                 relief='solid', bd=1,
                                 wrap=tk.WORD)          # 自動換行
        
        # 垂直捲軸
        scrollbar = tk.Scrollbar(text_frame, command=self.status_text.yview,
                               bg=self.colors['accent'])
        
        # 佈局文字框和捲軸
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.config(yscrollcommand=scrollbar.set)
        
        # === 顯示初始狀態訊息 ===
        if self.graphics_available:
            self.log_message("🎉 程式已啟動！圖形化功能已啟用，請選擇 CAD 檔案開始轉換流程。")
        else:
            self.log_message("🎉 程式已啟動！使用基本功能模式，請選擇 CAD 檔案開始轉換流程。")
        
        # === 設定鍵盤快捷鍵 ===
        self.setup_keyboard_shortcuts()
    
    def add_hover_effects(self):
        """
        添加按鈕滑鼠懸停效果
        
        功能:
        - 為按鈕添加滑鼠進入和離開時的視覺效果
        - 可擴展為其他元件添加互動效果
        
        注意: 目前在 create_material_button 方法中已實現
        """
        def on_enter(event, widget, hover_color):
            """滑鼠進入時改變背景色"""
            widget.configure(bg=hover_color)
        
        def on_leave(event, widget, normal_color):
            """滑鼠離開時恢復原背景色"""
            widget.configure(bg=normal_color)
        
        # 為所有按鈕添加 hover 效果（這裡可以進一步擴展）
        pass
    
    def setup_keyboard_shortcuts(self):
        """
        設定鍵盤快捷鍵
        
        快捷鍵列表:
        - Ctrl+O: 開啟 CAD 檔案
        - Ctrl+S: 設定 Excel 輸出檔案
        - F5: 開始轉換
        - Esc: 重置表單
        
        功能:
        - 提高使用者操作效率
        - 符合一般軟體的快捷鍵習慣
        """
        self.root.bind('<Control-o>', lambda e: self.browse_cad_file())    # Ctrl+O 開檔
        self.root.bind('<Control-s>', lambda e: self.browse_excel_file())  # Ctrl+S 存檔
        self.root.bind('<F5>', lambda e: self.start_conversion())          # F5 轉換
        self.root.bind('<Escape>', lambda e: self.reset_form())            # Esc 重置
    
    def show_file_info(self, file_path):
        """
        顯示選取檔案的基本資訊
        
        Args:
            file_path (str): 檔案完整路徑
        
        功能:
        - 顯示檔案大小
        - 快速檢查 DXF 檔案的文字實體數量
        - 提供檔案狀態回饋給使用者
        """
        try:
            # 檢查檔案是否存在
            if os.path.isfile(file_path):
                # 取得檔案大小
                file_size = os.path.getsize(file_path)
                size_text = self.format_file_size(file_size)
                info_text = f"📄 {size_text}"
                
                # 針對 DXF 檔案進行額外檢查
                if file_path.lower().endswith('.dxf'):
                    try:
                        # 快速讀取 DXF 檔案並計算文字實體數量
                        doc = ezdxf.readfile(file_path)
                        text_count = len(list(doc.modelspace().query("TEXT")))
                        info_text += f" | 📝 {text_count} 個文字實體"
                    except:
                        info_text += " | ⚠️ 檔案格式驗證中..."
                
                # 更新檔案資訊標籤
                self.file_info_label.config(text=info_text)
            else:
                self.file_info_label.config(text="❌ 檔案不存在")
                
        except Exception as e:
            self.file_info_label.config(text=f"❌ 讀取錯誤: {str(e)}")
    
    def format_file_size(self, size_bytes):
        """
        格式化檔案大小顯示
        
        Args:
            size_bytes (int): 檔案大小（位元組）
        
        Returns:
            str: 格式化後的檔案大小字串
        
        功能:
        - 自動選擇適當的單位（B, KB, MB, GB）
        - 保留適當的小數位數
        """
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024**2:
            return f"{size_bytes/1024:.1f} KB"
        elif size_bytes < 1024**3:
            return f"{size_bytes/(1024**2):.1f} MB"
        else:
            return f"{size_bytes/(1024**3):.1f} GB"
    
    def init_progress(self, steps_config):
        """
        初始化進度追蹤系統
        
        Args:
            steps_config (list): 處理步驟列表，每個元素為步驟描述
        
        功能:
        - 重設進度相關變數
        - 設定總步驟數和步驟描述
        - 記錄處理開始時間
        """
        self.current_step = 0                                          # 重設當前步驟
        self.total_steps = len(steps_config)                          # 設定總步驟數
        self.step_descriptions = {i: desc for i, desc in enumerate(steps_config)}  # 建立步驟描述字典
        self.processing_start_time = time.time()                      # 記錄開始時間
        self.progress["maximum"] = 100                                 # 設定進度條最大值
        self.progress["value"] = 0                                     # 重設進度條值
    
    def update_progress(self, step=None, detail="", percentage=None):
        """
        更新進度顯示
        
        Args:
            step (int, optional): 當前步驟編號
            detail (str): 詳細處理說明
            percentage (float, optional): 自訂進度百分比
        
        功能:
        - 更新進度條和步驟顯示
        - 計算並顯示處理時間和預估剩餘時間
        - 強制更新使用者介面以保持回應性
        """
        try:
            # === 更新步驟編號 ===
            if step is not None:
                self.current_step = step
            
            # === 計算進度百分比 ===
            if percentage is not None:
                progress_value = percentage
            else:
                # 根據當前步驟計算百分比
                progress_value = (self.current_step / self.total_steps) * 100 if self.total_steps > 0 else 0
            
            # === 更新進度條 ===
            self.progress["value"] = progress_value
            
            # === 更新當前步驟顯示 ===
            if self.current_step < len(self.step_descriptions):
                # 步驟圖示列表
                step_icons = ["🔍", "📂", "📊", "⚙️", "📈", "📋", "🎨", "💾"]
                # 選擇對應圖示，超出範圍時使用預設圖示
                icon = step_icons[self.current_step] if self.current_step < len(step_icons) else "⭐"
                # 組合步驟文字
                step_text = f"{icon} 步驟 {self.current_step + 1}/{self.total_steps}: {self.step_descriptions[self.current_step]}"
                self.current_step_label.config(text=step_text)
            
            # === 更新詳細進度說明 ===
            if detail:
                self.detail_progress_label.config(text=f"🔄 {detail}")
            
            # === 計算並顯示時間資訊 ===
            elapsed_time = time.time() - self.processing_start_time
            if progress_value > 0:
                # 預估總時間和剩餘時間
                estimated_total = elapsed_time * (100 / progress_value)
                remaining_time = estimated_total - elapsed_time
                time_text = f"⏱️ 已用時: {self.format_time(elapsed_time)} | ⏳ 預估剩餘: {self.format_time(remaining_time)} | 📊 {progress_value:.1f}%"
            else:
                time_text = f"⏱️ 已用時: {self.format_time(elapsed_time)} | 📊 0.0%"
            
            self.time_label.config(text=time_text)
            
            # === 強制更新使用者介面 ===
            self.root.update_idletasks()
            
        except Exception as e:
            print(f"更新進度時發生錯誤: {str(e)}")
    
    def format_time(self, seconds):
        """
        格式化時間顯示
        
        Args:
            seconds (float): 秒數
        
        Returns:
            str: 格式化後的時間字串
        
        功能:
        - 自動選擇適當的時間單位
        - 小於1分鐘顯示秒數
        - 小於1小時顯示分鐘和秒數
        - 超過1小時顯示小時、分鐘
        """
        if seconds < 60:
            return f"{seconds:.1f}秒"
        elif seconds < 3600:
            return f"{int(seconds // 60)}分{int(seconds % 60)}秒"
        else:
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            return f"{hours}小時{minutes}分"
    
    def browse_cad_file(self):
        """
        瀏覽並選擇 CAD 檔案
        
        功能:
        - 開啟檔案對話框供使用者選擇 CAD 檔案
        - 支援 DXF 和 DWG 檔案格式
        - 自動設定對應的 Excel 輸出檔案名稱
        - 顯示檔案基本資訊
        """
        # 定義支援的檔案類型
        filetypes = (
            ("DXF 檔案", "*.dxf"),
            ("DWG 檔案", "*.dwg"),
            ("所有檔案", "*.*")
        )
        
        # 開啟檔案選擇對話框
        filename = filedialog.askopenfilename(
            title="選擇 CAD 檔案",
            filetypes=filetypes
        )
        
        if filename:
            # 設定 CAD 檔案路徑
            self.cad_path.set(filename)
            
            # 自動生成 Excel 檔案路徑
            default_excel = os.path.splitext(filename)[0] + "_鋼筋計料.xlsx"
            self.excel_path.set(default_excel)
            
            # 顯示檔案資訊
            self.show_file_info(filename)
            
            # 記錄操作
            self.log_message(f"✅ 已選擇檔案: {os.path.basename(filename)}")
    
    def browse_excel_file(self):
        """
        瀏覽並設定 Excel 輸出檔案
        
        功能:
        - 開啟儲存檔案對話框
        - 自動添加 .xlsx 副檔名
        - 確認覆寫既有檔案
        """
        # 定義支援的檔案類型
        filetypes = (
            ("Excel 檔案", "*.xlsx"),
            ("所有檔案", "*.*")
        )
        
        # 開啟儲存檔案對話框
        filename = filedialog.asksaveasfilename(
            title="儲存 Excel 檔案",
            filetypes=filetypes,
            defaultextension=".xlsx",    # 預設副檔名
            confirmoverwrite=True        # 確認覆寫
        )
        
        if filename:
            # 設定 Excel 檔案路徑
            self.excel_path.set(filename)
            # 記錄操作
            self.log_message(f"📁 輸出路徑: {os.path.basename(filename)}")
    
    def reset_form(self):
        """
        重置表單和狀態
        
        功能:
        - 清空所有輸入欄位
        - 重設進度顯示
        - 恢復按鈕狀態
        - 清除狀態日誌
        """
        # === 清空輸入欄位 ===
        self.cad_path.set("")                           # 清空 CAD 檔案路徑
        self.excel_path.set("")                         # 清空 Excel 檔案路徑
        self.file_info_label.config(text="")           # 清空檔案資訊顯示
        
        # === 重設狀態顯示 ===
        self.status_text.delete(1.0, tk.END)           # 清空狀態日誌
        self.progress["value"] = 0                      # 重設進度條
        self.current_step_label.config(text="⭐ 準備就緒")  # 重設步驟顯示
        self.detail_progress_label.config(text="")     # 清空詳細進度
        self.time_label.config(text="")                # 清空時間顯示
        
        # === 恢復按鈕狀態 ===
        self.convert_button.config(state="normal")     # 啟用轉換按鈕
        
        # 根據圖形功能狀態設定按鈕文字
        if self.graphics_available:
            self.convert_button.config(text="🚀 開始轉換 (含圖形)")
        else:
            self.convert_button.config(text="🚀 開始轉換")
        
        # 記錄重置操作
        self.log_message("🔄 表單已重置，請重新選擇檔案。")
    
    def log_message(self, message):
        """
        記錄訊息到狀態文字框
        
        Args:
            message (str): 要記錄的訊息
        
        功能:
        - 自動添加時間戳記
        - 顯示在狀態文字框中
        - 自動捲動到最新訊息
        - 同時輸出到控制台
        """
        try:
            # 產生時間戳記
            timestamp = time.strftime("%H:%M:%S")
            formatted_message = f"[{timestamp}] {message}"
            
            # 確保訊息以換行結尾
            if not formatted_message.endswith('\n'):
                formatted_message += '\n'
            
            # 插入到狀態文字框末尾
            self.status_text.insert(tk.END, formatted_message)
            # 自動捲動到最新訊息
            self.status_text.see(tk.END)
            # 強制更新介面
            self.root.update_idletasks()
            
            # 同時輸出到控制台供除錯使用
            print(formatted_message.strip())
            
        except Exception as e:
            print(f"記錄訊息時發生錯誤: {str(e)}")
    
    def start_conversion(self):
        """
        開始 CAD 到 Excel 的轉換流程
        
        功能:
        - 驗證輸入檔案
        - 設定轉換狀態
        - 在新執行緒中執行轉換以避免 UI 凍結
        """
        # === 驗證 CAD 檔案 ===
        if not self.cad_path.get():
            messagebox.showerror("❌ 錯誤", "請先選擇 CAD 檔案！")
            return
        
        # === 驗證 Excel 輸出檔案 ===
        if not self.excel_path.get():
            messagebox.showerror("❌ 錯誤", "請先指定 Excel 輸出檔案！")
            return
        
        # === 準備轉換環境 ===
        # 清空狀態文字
        self.status_text.delete(1.0, tk.END)
        
        # 禁用轉換按鈕並更新狀態
        self.convert_button.config(state="disabled", text="🔄 轉換中...", bg=self.colors['text_secondary'])
        
        # === 在新執行緒中執行轉換 ===
        # 使用執行緒避免 UI 凍結，讓使用者可以看到即時進度
        conversion_thread = threading.Thread(target=self.convert_cad_to_excel)
        conversion_thread.daemon = True  # 設為守護執行緒，主程式結束時自動終止
        conversion_thread.start()
    
    # ====== 原有的計算方法 ======
    
    def calculate_line_length(self, start_point, end_point):
        """
        計算兩點之間的距離
        
        Args:
            start_point (tuple): 起點座標 (x, y, z)
            end_point (tuple): 終點座標 (x, y, z)
        
        Returns:
            float: 兩點間的距離
        
        功能:
        - 支援 2D 和 3D 座標計算
        - 使用歐幾里得距離公式
        """
        return math.sqrt(
            (end_point[0] - start_point[0])**2 + 
            (end_point[1] - start_point[1])**2 + 
            (end_point[2] - start_point[2])**2 if len(start_point) > 2 and len(end_point) > 2 
            else (end_point[2] if len(end_point) > 2 else 0)
        )
    
    def calculate_polyline_length(self, points):
        """
        計算多段線的總長度
        
        Args:
            points (list): 點座標列表
        
        Returns:
            float: 多段線的總長度
        
        功能:
        - 逐段計算相鄰點之間的距離
        - 累加得到總長度
        """
        total_length = 0
        for i in range(len(points) - 1):
            total_length += self.calculate_line_length(points[i], points[i+1])
        return total_length
    
    def extract_rebar_info(self, text):
        """
        從文字中提取鋼筋資訊
        
        Args:
            text (str): CAD 文字內容
        
        Returns:
            tuple: (鋼筋編號, 數量, 長度, 分段長度列表)
        
        功能:
        - 使用正規表達式識別鋼筋編號（#4, D13, 13mm 等格式）
        - 提取數量和長度資訊
        - 處理分段長度（如 150+200+150）
        - 支援多種 CAD 標記格式
        """
        if not text:
            return None, None, None, None
            
        number = ""      # 鋼筋編號
        count = 1        # 數量（預設為1）
        length = None    # 總長度
        segments = []    # 分段長度
        
        # === 尋找鋼筋號數（支援多種格式） ===
        
        # 格式1: #4, #5, #6 等
        number_match = re.search(r'#(\d+)', text)
        if number_match:
            number = "#" + number_match.group(1)
        
        # 格式2: D13, D16, D19 等（轉換為對應的 # 編號）
        if not number:
            number_match = re.search(r'D(\d+)', text)
            if number_match:
                diameter = float(number_match.group(1))
                # 根據直徑對應到標準鋼筋編號
                for num, dia in self.rebar_unit_weight.items():
                    if abs(dia - diameter) < 0.1:
                        number = num
                        break
        
        # 格式3: 13mm, 16mm 等（轉換為對應的 # 編號）
        if not number:
            number_match = re.search(r'(\d+(?:\.\d+)?)\s*mm', text)
            if number_match:
                diameter = float(number_match.group(1))
                # 根據直徑對應到標準鋼筋編號
                for num, dia in self.rebar_unit_weight.items():
                    if abs(dia - diameter) < 0.1:
                        number = num
                        break
        
        # === 尋找長度和數量 ===
        
        # 格式: #4-150+200+150x5（分段長度和數量）
        length_count_match = re.search(r'[#D]?\d+[-_](?:(\d+(?:\+\d+)*))[xX×*-](\d+)', text)
        if length_count_match:
            try:
                # 解析分段長度
                length_parts = length_count_match.group(1).split('+')
                segments = [float(part) for part in length_parts]
                total_length = sum(segments)
                length = total_length
                # 解析數量
                count = int(length_count_match.group(2))
            except:
                length = None
                count = 1
        else:
            # 簡單格式: 只有數量標記（如 x5, ×3）
            count_match = re.search(r'[xX×*-](\d+)', text)
            if count_match:
                try:
                    count = int(count_match.group(1))
                except:
                    count = 1
        
        return number, count, length, segments
    
    def get_rebar_diameter(self, number):
        """
        根據鋼筋號數獲取直徑
        
        Args:
            number (str): 鋼筋編號（如 "#4"）
        
        Returns:
            float: 鋼筋直徑（公釐）
        
        功能:
        - 提供標準鋼筋編號對應的直徑
        - 用於計算彎曲半徑等參數
        """
        rebar_diameter = {
            "#2": 6.4, "#3": 9.5, "#4": 12.7, "#5": 15.9, "#6": 19.1,
            "#7": 22.2, "#8": 25.4, "#9": 28.7, "#10": 32.3, "#11": 35.8,
            "#12": 43.0, "#13": 43.0, "#14": 43.0, "#15": 43.0, "#16": 43.0,
            "#17": 43.0, "#18": 57.3
        }
        return rebar_diameter.get(number, "")
    
    def get_rebar_unit_weight(self, number):
        """
        根據鋼筋號數獲取單位重量
        
        Args:
            number (str): 鋼筋編號（如 "#4"）
        
        Returns:
            float: 單位重量（kg/m）
        
        功能:
        - 提供標準鋼筋編號對應的單位重量
        - 用於計算總重量
        """
        return self.rebar_unit_weight.get(number, 0)
    
    def calculate_rebar_weight(self, number, length, count=1):
        """
        計算鋼筋總重量
        
        Args:
            number (str): 鋼筋編號
            length (float): 長度（公分）
            count (int): 數量
        
        Returns:
            float: 總重量（公斤）
        
        功能:
        - 根據鋼筋編號取得單位重量
        - 將長度單位從公分轉換為公尺
        - 計算總重量 = 單位重量 × 長度(m) × 數量
        """
        unit_weight = self.get_rebar_unit_weight(number)
        if unit_weight and length:
            length_m = length / 100.0  # 公分轉公尺
            return round(unit_weight * length_m * count, 2)
        return 0
    
    # ====== 保留原有的 ASCII 繪圖方法作為備用 ======
    
    def draw_ascii_rebar(self, segments):
        """
        使用 ASCII 字元繪製彎折示意圖（備用方法）
        
        Args:
            segments (list): 分段長度列表
        
        Returns:
            str: ASCII 字元組成的鋼筋圖示
        
        功能:
        - 當圖形功能不可用時的備用方案
        - 使用純文字字元繪製簡單的鋼筋形狀
        - 支援直線和彎曲形狀的基本表示
        """
        if not segments:
            return "─"
            
        # === 單段 = 直線 ===
        if len(segments) == 1:
            length = str(int(segments[0]))
            line = "─" * 10
            total_width = max(len(line), len(length))
            length_spaces = (total_width - len(length)) // 2
            line_spaces = (total_width - len(line)) // 2
            return f"{' ' * length_spaces}{length}\n{' ' * line_spaces}{line}"
        
        # === 多段 = 彎曲形狀 ===
        lines = []
        middle_chars = 10
        
        line = ""
        start_num = str(int(segments[0]))
        line += f"{start_num} |"
        
        if len(segments) > 1:
            line += "─" * middle_chars
        
        if len(segments) > 2:
            end_num = str(int(segments[-1]))
            line += f"| {end_num}"
        elif len(segments) == 2:
            end_num = str(int(segments[-1]))
            line += f"| {end_num}"
        
        if len(segments) > 1:
            middle_num = str(int(sum(segments[1:-1] if len(segments) > 2 else segments[1:])))
            total_width = len(line)
            start_pos = len(start_num) + 2
            middle_section_width = middle_chars
            middle_start = start_pos + (middle_section_width - len(middle_num)) // 2
            
            first_line = " " * total_width
            first_line = first_line[:middle_start] + middle_num + first_line[middle_start + len(middle_num):]
            lines.append(first_line)
        
        lines.append(line)
        return "\n".join(lines)
    
    # ====== 主要轉換方法 ======
    
    def convert_cad_to_excel(self):
        """
        CAD 到 Excel 的主要轉換流程
        
        功能:
        - 完整的轉換流程控制
        - 進度追蹤和錯誤處理
        - 生成專業的 Excel 報表
        
        處理步驟:
        1. 驗證檔案
        2. 載入 CAD 檔案
        3. 分析文字實體
        4. 處理鋼筋資料
        5. 生成統計資料
        6. 建立 Excel 工作簿
        7. 格式化 Excel
        8. 儲存檔案
        """
        try:
            # === 定義處理步驟 ===
            steps = [
                "驗證檔案",
                "載入 CAD 檔案", 
                "分析文字實體",
                "處理鋼筋資料",
                "生成統計資料",
                "建立 Excel 工作簿",
                "格式化 Excel",
                "儲存檔案"
            ]
            
            # 初始化進度追蹤
            self.init_progress(steps)
            self.log_message("🚀 開始轉換流程...")
            
            # === 步驟 1: 驗證檔案 ===
            self.update_progress(0, "正在驗證檔案...")
            self.log_message(f"📂 檢查檔案: {os.path.basename(self.cad_path.get())}")
            
            # 檢查檔案是否存在
            if not os.path.isfile(self.cad_path.get()):
                raise FileNotFoundError("找不到指定的 CAD 檔案")
                
            # 檢查檔案格式
            file_ext = os.path.splitext(self.cad_path.get())[1].lower()
            if file_ext != '.dxf':
                self.log_message("⚠️ 警告: 非 DXF 檔案可能無法正確讀取")
            
            # === 步驟 2: 載入 CAD 檔案 ===
            self.update_progress(1, "正在載入 CAD 檔案...")
            
            try:
                # 使用 ezdxf 讀取 DXF 檔案
                doc = ezdxf.readfile(self.cad_path.get())
                msp = doc.modelspace()  # 取得模型空間
                self.log_message("✅ CAD 檔案載入成功")
            except Exception as e:
                raise Exception(f"無法讀取 CAD 檔案: {str(e)}")
            
            # === 步驟 3: 分析文字實體 ===
            self.update_progress(2, "正在分析文字實體...")
            
            # 查詢所有文字實體
            text_entities = list(msp.query("TEXT"))
            entity_count = len(text_entities)
            self.log_message(f"📊 找到 {entity_count} 個文字實體")
            
            if entity_count == 0:
                raise Exception("未找到任何文字實體")
            
            # === 步驟 4: 處理鋼筋資料 ===
            self.update_progress(3, "正在處理鋼筋資料...")
            
            rebar_data = []           # 儲存鋼筋資料
            processed_count = 0       # 已處理數量
            valid_rebar_count = 0     # 有效鋼筋數量
            
            # 逐一處理每個文字實體
            for i, text in enumerate(text_entities):
                # 更新子進度（30-50%）
                sub_progress = 30 + (i / entity_count) * 20
                detail = f"處理文字實體 {i+1}/{entity_count}"
                self.update_progress(percentage=sub_progress, detail=detail)
                
                # 取得文字內容
                text_content = text.dxf.text
                if text_content:
                    # 提取鋼筋資訊
                    number, count, length, segments = self.extract_rebar_info(text_content)
                    
                    if number and length is not None:
                        valid_rebar_count += 1
                        
                        # === 計算相關參數 ===
                        length_cm = length
                        unit_weight = self.get_rebar_unit_weight(number) if number.startswith("#") else 0
                        weight = self.calculate_rebar_weight(number, length_cm, count) if number.startswith("#") else 0
                        material_grade = self._get_material_grade(number)
                        
                        # === 建立資料記錄 ===
                        data = {
                            "編號": number,
                            "長度(cm)": round(length_cm, 2),
                            "數量": count,
                            "單位重(kg/m)": unit_weight,
                            "總重量(kg)": weight,
                            "材料等級": material_grade,
                            "圖層": text.dxf.layer,
                            "備註": text_content
                        }
                        
                        # === 添加分段長度欄位 ===
                        if segments:
                            for j, segment in enumerate(segments):
                                letter = chr(65 + j)  # A, B, C, ...
                                data[f"{letter}(cm)"] = round(segment, 2)
                        
                        rebar_data.append(data)
                        
                        # 記錄前5個鋼筋的詳細資訊
                        if valid_rebar_count <= 5:
                            self.log_message(f"✅ 鋼筋 #{valid_rebar_count}: {number}, {length_cm}cm, {count}支, {material_grade}")
                
                processed_count += 1
            
            self.log_message(f"📈 處理完成: 找到 {valid_rebar_count} 個有效鋼筋標記")
            
            if not rebar_data:
                raise Exception("沒有找到任何可轉換的鋼筋數據")
            
            # === 步驟 5: 生成統計資料 ===
            self.update_progress(4, "正在生成統計資料...")
            
            # 計算基本統計
            total_quantity = sum(item["數量"] for item in rebar_data)
            total_length = sum(item["長度(cm)"] * item["數量"] for item in rebar_data)
            total_weight = sum(item["總重量(kg)"] for item in rebar_data)
            
            # 按號數統計
            rebar_types = {}
            for item in rebar_data:
                rebar_num = item["編號"]
                if rebar_num not in rebar_types:
                    rebar_types[rebar_num] = {"count": 0, "weight": 0}
                rebar_types[rebar_num]["count"] += item["數量"]
                rebar_types[rebar_num]["weight"] += item["總重量(kg)"]
            
            self.log_message(f"📊 統計結果: 總數量 {total_quantity}支, 總重量 {total_weight:.2f}kg")
            self.log_message(f"🔧 鋼筋類型: {len(rebar_types)} 種")
            
            # 根據號數排序資料
            sorted_data = sorted(rebar_data, key=lambda x: x["編號"] if "#" in x["編號"] else "z" + x["編號"])
            
            # === 步驟 6: 建立 Excel 工作簿 ===
            self.update_progress(5, "正在建立 Excel 工作簿...")
            
            # === 重新排列欄位順序 ===
            base_columns = ["編號", "長度(cm)", "數量", "總重量(kg)", "材料等級", "鋼筋圖示", "備註"]
            
            # 找出所有分段長度欄位
            segment_columns = set()
            for item in sorted_data:
                for key in item.keys():
                    if key.endswith("(cm)") and key != "長度(cm)":
                        segment_columns.add(key)
            
            # 按字母順序排序分段長度欄位
            segment_columns = sorted(segment_columns)
            
            # 組合最終欄位順序：基本欄位 + 分段長度欄位
            columns = base_columns[:2] + segment_columns + base_columns[2:]
            
            # 建立 DataFrame
            df = pd.DataFrame(sorted_data)
            
            # === 添加鋼筋圖示欄位 ===
            self.update_progress(percentage=75, detail="正在生成專業鋼筋圖示...")
            
            if self.graphics_available:
                self.log_message("🎨 使用圖形化鋼筋圖示生成...")
            else:
                self.log_message("📝 使用基本 ASCII 圖示生成...")
            
            # 為每筆資料生成鋼筋圖示
            for index, row in df.iterrows():
                # 收集分段長度
                segments = []
                for key in sorted([k for k in row.keys() if k.endswith("(cm)") and k != "長度(cm)"]):
                    if not pd.isna(row[key]):
                        segments.append(row[key])
                
                # 生成圖示
                df.at[index, "鋼筋圖示"] = self.enhanced_draw_rebar_diagram(segments, row["編號"])
            
            # 重新排列 DataFrame 欄位
            df = df.reindex(columns=columns)
            
            # === 步驟 7: 格式化 Excel ===
            self.update_progress(6, "正在格式化 Excel...")
            
            try:
                # === 創建 Excel 工作簿 ===
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "鋼筋計料表"
                
                # === 設定標題區域 ===
                project_name = os.path.basename(self.cad_path.get())
                graphics_note = "（含專業圖形化圖示）" if self.graphics_available else "（基本模式）"
                
                # 標題內容
                ws['A1'] = f"🏗️ 專業鋼筋計料表 {graphics_note}"
                ws['A2'] = f"專案名稱: {project_name}"
                ws['A3'] = f"生成時間: {time.strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A4'] = f"工具版本: CAD 鋼筋計料轉換工具 Pro v2.1"
                
                # 合併儲存格
                ws.merge_cells('A1:H1')
                ws.merge_cells('A2:H2')
                ws.merge_cells('A3:H3')
                ws.merge_cells('A4:H4')
                
                # === 設定標題樣式 ===
                title_font = Font(bold=True, size=16)
                subtitle_font = Font(bold=True, size=12)
                info_font = Font(size=10)
                
                ws['A1'].font = title_font
                ws['A2'].font = subtitle_font
                ws['A3'].font = info_font
                ws['A4'].font = info_font
                
                # 居中對齊
                title_align = Alignment(horizontal='center', vertical='center')
                ws['A1'].alignment = title_align
                ws['A2'].alignment = title_align
                ws['A3'].alignment = title_align
                ws['A4'].alignment = title_align
                
                # === 設定表頭（從第6行開始） ===
                headers = columns
                header_row = 6
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col_num)
                    cell.value = header
                    
                    # 表頭樣式
                    header_font = Font(bold=True)
                    header_align = Alignment(horizontal='center', vertical='center')
                    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                    header_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.fill = header_fill
                    cell.border = header_border
                
                # === 寫入資料（從第7行開始） ===
                data_start_row = 7
                row_num = data_start_row
                
                self.update_progress(percentage=80, detail="正在寫入資料...")
                
                for i, (_, row) in enumerate(df.iterrows()):
                    # 更新子進度（80-90%）
                    if len(df) > 0:
                        sub_progress = 80 + (i / len(df)) * 10
                        if i % 10 == 0:  # 每10行更新一次進度
                            self.update_progress(percentage=sub_progress, detail=f"寫入資料行 {i+1}/{len(df)}")
                    
                    # 寫入每個欄位的值
                    for col_num, col_name in enumerate(headers, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        
                        # 設定儲存格值
                        if col_name in row:
                            cell.value = row[col_name]
                        
                        # === 設定資料樣式 ===
                        data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        data_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        cell.alignment = data_align
                        cell.border = data_border
                        
                        # === 根據內容設定特殊樣式 ===
                        # 鋼筋編號欄位使用粗體
                        if col_num == 1 and row["編號"].startswith("#"):
                            cell.font = Font(bold=True)
                        
                        # 根據材料等級設定背景顏色
                        if col_name == "材料等級":
                            grade = row.get("材料等級", "")
                            if grade == "SD280":
                                cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                            elif grade == "SD420":
                                cell.fill = PatternFill(start_color="FFF2E7", end_color="FFF2E7", fill_type="solid")
                            elif grade == "SD490":
                                cell.fill = PatternFill(start_color="F3E7FF", end_color="F3E7FF", fill_type="solid")
                    
                    # 調整行高以適應圖示內容
                    ws.row_dimensions[row_num].height = 120 if self.graphics_available else 80
                    row_num += 1
                
                # === 添加統計總計行 ===
                summary_row = row_num + 1
                ws.cell(row=summary_row, column=1).value = "總計"
                ws.cell(row=summary_row, column=1).font = Font(bold=True)
                
                # 計算並填入總數量
                quantity_col = headers.index("數量") + 1
                ws.cell(row=summary_row, column=quantity_col).value = df["數量"].sum()
                ws.cell(row=summary_row, column=quantity_col).font = Font(bold=True)
                
                # 計算並填入總重量
                weight_col = headers.index("總重量(kg)") + 1
                ws.cell(row=summary_row, column=weight_col).value = round(df["總重量(kg)"].sum(), 2)
                ws.cell(row=summary_row, column=weight_col).font = Font(bold=True)
                
                # === 為統計行設定樣式 ===
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=summary_row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='double')  # 底部使用雙線
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                
                # === 設定欄位寬度 ===
                column_widths = {
                    "編號": 12,
                    "長度(cm)": 15,
                    "數量": 10,
                    "總重量(kg)": 15,
                    "材料等級": 15,
                    "鋼筋圖示": 80 if self.graphics_available else 40,  # 圖形模式需要更寬
                    "備註": 30
                }
                
                # 設定分段長度欄位的寬度
                for col in segment_columns:
                    column_widths[col] = 12
                
                # 應用欄位寬度設定
                for col_num, header in enumerate(headers, 1):
                    if header in column_widths:
                        column_letter = openpyxl.utils.get_column_letter(col_num)
                        ws.column_dimensions[column_letter].width = column_widths[header]
                
                # === 添加材料等級說明區域 ===
                legend_start_row = summary_row + 3
                ws.cell(row=legend_start_row, column=1).value = "材料等級說明:"
                ws.cell(row=legend_start_row, column=1).font = Font(bold=True)
                
                # 材料等級對照表
                legend_data = [
                    ("SD280", "280 MPa", "一般結構用鋼筋"),
                    ("SD420", "420 MPa", "高強度結構鋼筋"),
                    ("SD490", "490 MPa", "特殊高強度鋼筋")
                ]
                
                # 填入材料等級說明
                for i, (grade, strength, desc) in enumerate(legend_data):
                    row = legend_start_row + 1 + i
                    ws.cell(row=row, column=1).value = grade
                    ws.cell(row=row, column=2).value = strength
                    ws.cell(row=row, column=3).value = desc
                    
                    # 設定說明區域的樣式
                    for col in range(1, 4):
                        cell = ws.cell(row=row, column=col)
                        cell.font = Font(size=9)
                        # 根據等級設定對應的背景色
                        if grade == "SD280":
                            cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                        elif grade == "SD420":
                            cell.fill = PatternFill(start_color="FFF2E7", end_color="FFF2E7", fill_type="solid")
                        elif grade == "SD490":
                            cell.fill = PatternFill(start_color="F3E7FF", end_color="F3E7FF", fill_type="solid")
                
                self.log_message("🎨 Excel 格式設定完成")
                
            except Exception as e:
                # 格式化過程出錯時的處理
                self.log_message(f"⚠️ Excel 格式化時出現問題: {str(e)}")
                self.log_message("🔄 使用基本儲存方式...")
                # 如果格式化失敗，至少確保數據被保存
                df.to_excel(self.excel_path.get(), sheet_name='鋼筋計料', index=False)
            
            # === 步驟 8: 儲存檔案 ===
            self.update_progress(7, "正在儲存檔案...")
            
            try:
                wb.save(self.excel_path.get())
                self.log_message(f"💾 檔案已儲存: {os.path.basename(self.excel_path.get())}")
            except Exception as e:
                raise Exception(f"儲存檔案時發生錯誤: {str(e)}")
            
            # === 轉換完成 ===
            self.update_progress(percentage=100, detail="轉換完成!")
            
            # 計算處理時間
            elapsed_time = time.time() - self.processing_start_time
            
            # === 顯示完成統計 ===
            self.log_message("=" * 60)
            self.log_message("🎉 轉換完成!")
            self.log_message(f"📊 處理統計:")
            self.log_message(f"   • 處理檔案: {os.path.basename(self.cad_path.get())}")
            self.log_message(f"   • 文字實體: {entity_count} 個")
            self.log_message(f"   • 有效鋼筋: {valid_rebar_count} 個")
            self.log_message(f"   • 鋼筋類型: {len(rebar_types)} 種")
            self.log_message(f"   • 總數量: {total_quantity} 支")
            self.log_message(f"   • 總長度: {total_length:.2f} cm")
            self.log_message(f"   • 總重量: {total_weight:.2f} kg")
            self.log_message(f"   • 處理時間: {self.format_time(elapsed_time)}")
            self.log_message(f"   • 圖示類型: {'專業圖形化' if self.graphics_available else '基本ASCII'}")
            self.log_message(f"   • 輸出檔案: {os.path.basename(self.excel_path.get())}")
            self.log_message("=" * 60)
            
            # === 恢復按鈕狀態 ===
            if self.graphics_available:
                self.convert_button.config(state="normal", text="🚀 開始轉換 (含圖形)", bg=self.colors['success'])
            else:
                self.convert_button.config(state="normal", text="🚀 開始轉換", bg=self.colors['success'])
            
            # === 顯示完成對話框 ===
            graphics_note = "包含專業圖形化鋼筋圖示" if self.graphics_available else "使用基本圖示模式"
            result_message = f"""🎉 轉換完成！

📊 處理結果:
• 有效鋼筋: {valid_rebar_count} 個
• 鋼筋類型: {len(rebar_types)} 種  
• 總數量: {total_quantity} 支
• 總重量: {total_weight:.2f} kg
• 處理時間: {self.format_time(elapsed_time)}
• 圖示模式: {graphics_note}

💾 檔案已儲存至:
{self.excel_path.get()}

🎨 新功能特色:
✅ 自動材料等級識別 (SD280/SD420/SD490)
✅ 專業鋼筋彎曲形狀圖示
✅ 增強版 Excel 格式化
✅ 智能圖形降級機制

感謝您使用 CAD 鋼筋計料轉換工具 Pro v2.1！"""
            
            messagebox.showinfo("✅ 轉換完成", result_message)
            
        except Exception as e:
            # === 錯誤處理 ===
            self.log_message(f"❌ 錯誤: {str(e)}")
            self.update_progress(percentage=0, detail="轉換失敗")
            
            # 恢復按鈕狀態
            if self.graphics_available:
                self.convert_button.config(state="normal", text="🚀 開始轉換 (含圖形)", bg=self.colors['success'])
            else:
                self.convert_button.config(state="normal", text="🚀 開始轉換", bg=self.colors['success'])
            
            # 顯示錯誤對話框
            messagebox.showerror("❌ 轉換錯誤", f"轉換過程中發生錯誤:\n\n{str(e)}\n\n請檢查檔案格式和內容後重試。")


def main():
    """
    主程式入口點
    
    功能:
    - 建立 tkinter 主視窗
    - 初始化轉換器應用程式
    - 設定視窗屬性和事件處理
    - 啟動主要事件迴圈
    """
    # === 建立主視窗 ===
    root = tk.Tk()
    
    # === 設定視窗圖示（可選） ===
    try:
        # 如果有圖示檔案可以取消註解
        # root.iconbitmap('icon.ico')
        pass
    except:
        pass
    
    # === 視窗居中函數 ===
    def center_window(window, width=900, height=750):
        """
        將視窗置於螢幕中央
        
        Args:
            window: tkinter 視窗物件
            width (int): 視窗寬度
            height (int): 視窗高度
        """
        # 取得螢幕尺寸
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        
        # 計算置中位置
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        # 設定視窗位置和大小
        window.geometry(f"{width}x{height}+{x}+{y}")
    
    # 將視窗置中
    center_window(root)
    
    # === 建立應用程式實例 ===
    app = ModernCADtoExcelConverter(root)
    
    # === 設定視窗關閉事件 ===
    def on_closing():
        """
        視窗關閉時的確認對話框
        
        功能:
        - 詢問使用者是否確定退出
        - 避免意外關閉程式
        """
        if messagebox.askokcancel("退出確認", "確定要退出 CAD 鋼筋計料轉換工具嗎？"):
            root.destroy()
    
    # 綁定視窗關閉事件
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # === 啟動主要事件迴圈 ===
    root.mainloop()


# === 程式進入點 ===
if __name__ == "__main__":
    """
    當檔案被直接執行時啟動主程式
    
    用途:
    - 防止在被其他模組匯入時自動執行
    - 確保只有在直接執行此檔案時才啟動 GUI
    """
    main()